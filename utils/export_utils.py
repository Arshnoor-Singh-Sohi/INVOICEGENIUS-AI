"""
Export Utilities for InvoiceGenius AI
===================================

This module handles exporting processed invoice data into various professional formats.
Think of this as your personal report generator that can create beautiful, formatted
outputs of your invoice data for different business needs.

Why multiple export formats?
- Excel: Perfect for financial analysis, pivot tables, and business users
- PDF: Professional reports for presentations and archival
- JSON: Machine-readable format for API integrations
- CSV: Universal format that works with any spreadsheet software

Each export format is optimized for its specific use case, with proper formatting,
charts, and business-relevant insights included where appropriate.
"""

import os
import json
import csv
import logging
from datetime import datetime, date
from typing import Dict, List, Optional, Any
from pathlib import Path
import io

# Data processing
import pandas as pd

# Excel handling
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, BarChart, LineChart, Reference

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart

# Our custom modules
from config import Config

logger = logging.getLogger(__name__)

class ExportManager:
    """
    Professional export and reporting engine
    
    This class transforms raw invoice data into polished, business-ready reports
    and exports. It's designed to create outputs that look professional and
    provide real business value.
    
    Think of this as your personal assistant who can take a pile of invoice data
    and turn it into beautiful, insightful reports for executives, accountants,
    or business analysts.
    """
    
    def __init__(self):
        """Initialize the export manager with configuration"""
        self.config = Config()
        
        # Ensure exports directory exists
        self.config.EXPORTS_DIR.mkdir(exist_ok=True)
        
        # Set up styles for consistent formatting
        self._setup_styles()
        
        logger.info("Export manager initialized")
    
    def _setup_styles(self):
        """
        Set up consistent styling for all export formats
        
        This ensures all our exports have a professional, consistent look
        that reflects well on our application and the user's business.
        """
        # Excel styles
        self.excel_styles = {
            'header': {
                'font': Font(bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'currency': {
                'number_format': '$#,##0.00'
            },
            'date': {
                'number_format': 'yyyy-mm-dd'
            },
            'percentage': {
                'number_format': '0.0%'
            }
        }
        
        # PDF styles
        self.pdf_styles = getSampleStyleSheet()
        self.pdf_styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.pdf_styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#366092')
        ))
        
        # Company branding colors
        self.brand_colors = {
            'primary': '#366092',
            'secondary': '#52b788',
            'accent': '#f77f00',
            'text': '#2d3748',
            'light_gray': '#f7fafc'
        }
    
    def export_to_excel(self, invoice_data: List[Dict], filename: str = None) -> str:
        """
        Export invoice data to a professionally formatted Excel file
        
        Creates a comprehensive Excel workbook with multiple sheets:
        - Summary sheet with key metrics and charts
        - Detailed invoice data
        - Vendor analysis
        - Monthly trends
        
        Args:
            invoice_data: List of invoice dictionaries
            filename: Optional custom filename
            
        Returns:
            Path to the created Excel file
        """
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"invoice_export_{timestamp}.xlsx"
            
            filepath = self.config.EXPORTS_DIR / filename
            
            # Create DataFrame from invoice data
            df = self._prepare_dataframe(invoice_data)
            
            # Create Excel workbook with multiple sheets
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Sheet 1: Summary Dashboard
                self._create_summary_sheet(writer, df, invoice_data)
                
                # Sheet 2: Detailed Invoice Data
                self._create_detailed_sheet(writer, df)
                
                # Sheet 3: Vendor Analysis
                self._create_vendor_analysis_sheet(writer, df)
                
                # Sheet 4: Monthly Trends
                self._create_trends_sheet(writer, df)
                
                # Sheet 5: Raw Data (for power users)
                df.to_excel(writer, sheet_name='Raw Data', index=False)
            
            # Apply professional formatting
            self._format_excel_workbook(filepath)
            
            logger.info(f"Excel export created: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Failed to create Excel export: {str(e)}")
            raise
    
    def _prepare_dataframe(self, invoice_data: List[Dict]) -> pd.DataFrame:
        """
        Convert invoice data to pandas DataFrame with proper data types
        
        This ensures consistent data handling and enables powerful
        pandas operations for analysis and export.
        """
        # Flatten the data for DataFrame creation
        flattened_data = []
        
        for invoice in invoice_data:
            row = {
                'Invoice Number': invoice.get('invoice_number', ''),
                'Vendor Name': invoice.get('vendor_name', ''),
                'Invoice Date': self._parse_date(invoice.get('invoice_date')),
                'Due Date': self._parse_date(invoice.get('due_date')),
                'Total Amount': self._safe_float(invoice.get('total_amount')),
                'Subtotal': self._safe_float(invoice.get('subtotal')),
                'Tax Amount': self._safe_float(invoice.get('tax_amount')),
                'Currency': invoice.get('currency', 'USD'),
                'Payment Terms': invoice.get('payment_terms', ''),
                'PO Number': invoice.get('po_number', ''),
                'Confidence Score': self._safe_float(invoice.get('confidence')),
                'Validation Score': self._safe_float(invoice.get('validation_score')),
                'Processing Time': self._safe_float(invoice.get('processing_time')),
                'File Name': invoice.get('file_name', ''),
                'Processed At': self._parse_datetime(invoice.get('processed_at')),
                'AI Model': invoice.get('ai_model', '')
            }
            flattened_data.append(row)
        
        df = pd.DataFrame(flattened_data)
        
        # Convert data types
        if not df.empty:
            df['Invoice Date'] = pd.to_datetime(df['Invoice Date'])
            df['Due Date'] = pd.to_datetime(df['Due Date'])
            df['Processed At'] = pd.to_datetime(df['Processed At'])
        
        return df
    
    def _create_summary_sheet(self, writer, df: pd.DataFrame, raw_data: List[Dict]):
        """Create an executive summary sheet with key metrics and insights"""
        summary_data = []
        
        if not df.empty:
            # Key metrics
            summary_data.extend([
                ['Total Invoices', len(df)],
                ['Total Amount', f"${df['Total Amount'].sum():,.2f}"],
                ['Average Amount', f"${df['Total Amount'].mean():,.2f}"],
                ['Date Range', f"{df['Invoice Date'].min().strftime('%Y-%m-%d')} to {df['Invoice Date'].max().strftime('%Y-%m-%d')}"],
                ['Unique Vendors', df['Vendor Name'].nunique()],
                ['Average Confidence', f"{df['Confidence Score'].mean():.1%}"],
                ['Average Processing Time', f"{df['Processing Time'].mean():.2f}s"],
                ['']  # Empty row for spacing
            ])
            
            # Top vendors by amount
            vendor_totals = df.groupby('Vendor Name')['Total Amount'].sum().sort_values(ascending=False).head(5)
            summary_data.append(['Top 5 Vendors by Amount', ''])
            for vendor, amount in vendor_totals.items():
                summary_data.append([vendor, f"${amount:,.2f}"])
        
        summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _create_detailed_sheet(self, writer, df: pd.DataFrame):
        """Create detailed invoice listing with all fields"""
        df.to_excel(writer, sheet_name='Invoice Details', index=False)
    
    def _create_vendor_analysis_sheet(self, writer, df: pd.DataFrame):
        """Create vendor analysis with totals, averages, and frequency"""
        if df.empty:
            return
        
        vendor_analysis = df.groupby('Vendor Name').agg({
            'Total Amount': ['sum', 'mean', 'count'],
            'Invoice Date': ['min', 'max'],
            'Confidence Score': 'mean'
        }).round(2)
        
        # Flatten column names
        vendor_analysis.columns = ['Total Amount', 'Average Amount', 'Invoice Count', 
                                 'First Invoice', 'Last Invoice', 'Avg Confidence']
        
        vendor_analysis = vendor_analysis.sort_values('Total Amount', ascending=False)
        vendor_analysis.to_excel(writer, sheet_name='Vendor Analysis')
    
    def _create_trends_sheet(self, writer, df: pd.DataFrame):
        """Create monthly trends analysis"""
        if df.empty:
            return
        
        # Group by month
        df['Month'] = df['Invoice Date'].dt.to_period('M')
        monthly_trends = df.groupby('Month').agg({
            'Total Amount': 'sum',
            'Invoice Number': 'count',
            'Confidence Score': 'mean'
        }).round(2)
        
        monthly_trends.columns = ['Total Amount', 'Invoice Count', 'Avg Confidence']
        monthly_trends.to_excel(writer, sheet_name='Monthly Trends')
    
    def _format_excel_workbook(self, filepath: str):
        """Apply professional formatting to the Excel workbook"""
        try:
            workbook = openpyxl.load_workbook(filepath)
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Apply header styling
                if worksheet.max_row > 0:
                    for cell in worksheet[1]:  # First row
                        for style_key, style_value in self.excel_styles['header'].items():
                            setattr(cell, style_key, style_value)
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(filepath)
            
        except Exception as e:
            logger.warning(f"Failed to apply Excel formatting: {str(e)}")
    
    def export_to_pdf_report(self, invoice_data: List[Dict], filename: str = None) -> str:
        """
        Create a comprehensive PDF report with charts and analysis
        
        This creates a professional business report that executives and
        stakeholders can easily read and understand.
        """
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"invoice_report_{timestamp}.pdf"
            
            filepath = self.config.EXPORTS_DIR / filename
            
            # Create PDF document
            doc = SimpleDocTemplate(str(filepath), pagesize=A4)
            elements = []
            
            # Title page
            elements.extend(self._create_pdf_title_page(invoice_data))
            
            # Executive summary
            elements.extend(self._create_pdf_executive_summary(invoice_data))
            
            # Detailed analysis
            elements.extend(self._create_pdf_detailed_analysis(invoice_data))
            
            # Invoice listing
            elements.extend(self._create_pdf_invoice_listing(invoice_data))
            
            # Build PDF
            doc.build(elements)
            
            logger.info(f"PDF report created: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Failed to create PDF report: {str(e)}")
            raise
    
    def _create_pdf_title_page(self, invoice_data: List[Dict]) -> List:
        """Create an attractive title page for the PDF report"""
        elements = []
        
        # Title
        title = Paragraph(
            "Invoice Processing Report", 
            self.pdf_styles['CustomTitle']
        )
        elements.append(title)
        elements.append(Spacer(1, 0.5*inch))
        
        # Subtitle with date range
        if invoice_data:
            dates = [inv.get('invoice_date') for inv in invoice_data if inv.get('invoice_date')]
            if dates:
                min_date = min(dates)
                max_date = max(dates)
                subtitle = Paragraph(
                    f"Analysis Period: {min_date} to {max_date}",
                    self.pdf_styles['Heading2']
                )
                elements.append(subtitle)
        
        elements.append(Spacer(1, 0.3*inch))
        
        # Report metadata
        metadata = [
            f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
            f"Total Invoices: {len(invoice_data)}",
            f"Generated by: InvoiceGenius AI v{self.config.APP_VERSION}"
        ]
        
        for item in metadata:
            elements.append(Paragraph(item, self.pdf_styles['Normal']))
        
        elements.append(Spacer(1, 1*inch))
        
        return elements
    
    def _create_pdf_executive_summary(self, invoice_data: List[Dict]) -> List:
        """Create executive summary section"""
        elements = []
        
        elements.append(Paragraph("Executive Summary", self.pdf_styles['Heading1']))
        elements.append(Spacer(1, 0.2*inch))
        
        if not invoice_data:
            elements.append(Paragraph("No invoice data available for analysis.", self.pdf_styles['Normal']))
            return elements
        
        # Calculate key metrics
        total_amount = sum(self._safe_float(inv.get('total_amount', 0)) for inv in invoice_data)
        avg_amount = total_amount / len(invoice_data) if invoice_data else 0
        vendors = set(inv.get('vendor_name') for inv in invoice_data if inv.get('vendor_name'))
        avg_confidence = sum(self._safe_float(inv.get('confidence', 0)) for inv in invoice_data) / len(invoice_data)
        
        # Summary table
        summary_data = [
            ['Metric', 'Value'],
            ['Total Invoices Processed', f"{len(invoice_data):,}"],
            ['Total Invoice Amount', f"${total_amount:,.2f}"],
            ['Average Invoice Amount', f"${avg_amount:,.2f}"],
            ['Unique Vendors', f"{len(vendors):,}"],
            ['Average AI Confidence', f"{avg_confidence:.1%}"],
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.brand_colors['primary'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 0.5*inch))
        
        return elements
    
    def _create_pdf_detailed_analysis(self, invoice_data: List[Dict]) -> List:
        """Create detailed analysis section with insights"""
        elements = []
        
        elements.append(Paragraph("Detailed Analysis", self.pdf_styles['Heading1']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Vendor analysis
        vendor_totals = {}
        for inv in invoice_data:
            vendor = inv.get('vendor_name')
            amount = self._safe_float(inv.get('total_amount', 0))
            if vendor and amount:
                vendor_totals[vendor] = vendor_totals.get(vendor, 0) + amount
        
        if vendor_totals:
            elements.append(Paragraph("Top 5 Vendors by Total Amount", self.pdf_styles['Heading2']))
            
            # Sort vendors by total amount
            top_vendors = sorted(vendor_totals.items(), key=lambda x: x[1], reverse=True)[:5]
            
            vendor_data = [['Vendor', 'Total Amount', 'Percentage']]
            total_all_vendors = sum(vendor_totals.values())
            
            for vendor, amount in top_vendors:
                percentage = (amount / total_all_vendors) * 100 if total_all_vendors > 0 else 0
                vendor_data.append([vendor, f"${amount:,.2f}", f"{percentage:.1f}%"])
            
            vendor_table = Table(vendor_data)
            vendor_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.brand_colors['secondary'])),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(vendor_table)
        
        elements.append(Spacer(1, 0.5*inch))
        
        return elements
    
    def _create_pdf_invoice_listing(self, invoice_data: List[Dict]) -> List:
        """Create detailed invoice listing"""
        elements = []
        
        elements.append(Paragraph("Invoice Details", self.pdf_styles['Heading1']))
        elements.append(Spacer(1, 0.2*inch))
        
        if not invoice_data:
            elements.append(Paragraph("No invoices to display.", self.pdf_styles['Normal']))
            return elements
        
        # Create table with key invoice information
        table_data = [['Invoice #', 'Vendor', 'Date', 'Amount', 'Status']]
        
        for inv in invoice_data[:20]:  # Limit to first 20 for readability
            status = "✓ Processed" if self._safe_float(inv.get('confidence', 0)) > 0.8 else "⚠ Review"
            table_data.append([
                inv.get('invoice_number', 'N/A'),
                inv.get('vendor_name', 'N/A')[:25],  # Truncate long names
                inv.get('invoice_date', 'N/A'),
                f"${self._safe_float(inv.get('total_amount', 0)):,.2f}",
                status
            ])
        
        invoice_table = Table(table_data)
        invoice_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.brand_colors['primary'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(invoice_table)
        
        if len(invoice_data) > 20:
            elements.append(Spacer(1, 0.2*inch))
            elements.append(Paragraph(
                f"Note: Showing first 20 of {len(invoice_data)} total invoices. "
                "Full details available in Excel export.",
                self.pdf_styles['Normal']
            ))
        
        return elements
    
    def export_to_json(self, invoice_data: List[Dict], filename: str = None) -> str:
        """
        Export invoice data to structured JSON format
        
        Perfect for API integrations, data backups, or feeding data
        into other systems that can consume JSON.
        """
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"invoices_data_{timestamp}.json"
            
            filepath = self.config.EXPORTS_DIR / filename
            
            # Prepare data with metadata
            export_data = {
                'export_info': {
                    'generated_at': datetime.now().isoformat(),
                    'generator': f"InvoiceGenius AI v{self.config.APP_VERSION}",
                    'total_invoices': len(invoice_data),
                    'format_version': '1.0'
                },
                'invoices': invoice_data
            }
            
            # Write JSON with proper formatting
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
            
            logger.info(f"JSON export created: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Failed to create JSON export: {str(e)}")
            raise
    
    def export_to_csv(self, invoice_data: List[Dict], filename: str = None) -> str:
        """
        Export invoice data to CSV format
        
        Simple, universal format that works with any spreadsheet application.
        """
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"invoices_export_{timestamp}.csv"
            
            filepath = self.config.EXPORTS_DIR / filename
            
            # Convert to DataFrame and export
            df = self._prepare_dataframe(invoice_data)
            df.to_csv(filepath, index=False, encoding='utf-8')
            
            logger.info(f"CSV export created: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Failed to create CSV export: {str(e)}")
            raise
    
    def create_vendor_report(self, invoice_data: List[Dict], vendor_name: str) -> str:
        """
        Create a specialized report for a specific vendor
        
        This creates a focused analysis of all invoices from a particular vendor,
        useful for vendor relationship management and payment analysis.
        """
        try:
            # Filter data for specific vendor
            vendor_invoices = [
                inv for inv in invoice_data 
                if inv.get('vendor_name', '').lower() == vendor_name.lower()
            ]
            
            if not vendor_invoices:
                raise ValueError(f"No invoices found for vendor: {vendor_name}")
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"vendor_report_{vendor_name.replace(' ', '_')}_{timestamp}.pdf"
            filepath = self.config.EXPORTS_DIR / filename
            
            # Create vendor-specific PDF report
            doc = SimpleDocTemplate(str(filepath), pagesize=A4)
            elements = []
            
            # Title
            title = Paragraph(f"Vendor Report: {vendor_name}", self.pdf_styles['CustomTitle'])
            elements.append(title)
            elements.append(Spacer(1, 0.5*inch))
            
            # Vendor summary
            total_amount = sum(self._safe_float(inv.get('total_amount', 0)) for inv in vendor_invoices)
            avg_amount = total_amount / len(vendor_invoices)
            
            summary_data = [
                ['Metric', 'Value'],
                ['Total Invoices', f"{len(vendor_invoices):,}"],
                ['Total Amount', f"${total_amount:,.2f}"],
                ['Average Invoice Amount', f"${avg_amount:,.2f}"],
                ['Date Range', f"{min(inv.get('invoice_date', '') for inv in vendor_invoices)} to {max(inv.get('invoice_date', '') for inv in vendor_invoices)}"]
            ]
            
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.brand_colors['primary'])),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(summary_table)
            elements.append(Spacer(1, 0.5*inch))
            
            # Invoice details
            elements.append(Paragraph("Invoice Details", self.pdf_styles['Heading2']))
            
            invoice_details = [['Invoice #', 'Date', 'Amount', 'Due Date', 'PO Number']]
            for inv in vendor_invoices:
                invoice_details.append([
                    inv.get('invoice_number', 'N/A'),
                    inv.get('invoice_date', 'N/A'),
                    f"${self._safe_float(inv.get('total_amount', 0)):,.2f}",
                    inv.get('due_date', 'N/A'),
                    inv.get('po_number', 'N/A')
                ])
            
            details_table = Table(invoice_details)
            details_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(details_table)
            
            # Build PDF
            doc.build(elements)
            
            logger.info(f"Vendor report created: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Failed to create vendor report: {str(e)}")
            raise
    
    # Utility methods
    
    def _safe_float(self, value: Any) -> float:
        """Safely convert value to float"""
        try:
            return float(value) if value is not None else 0.0
        except (ValueError, TypeError):
            return 0.0
    
    def _parse_date(self, date_str: str) -> Optional[str]:
        """Parse date string to standard format"""
        if not date_str:
            return None
        
        try:
            # If it's already a date object, convert to string
            if isinstance(date_str, date):
                return date_str.strftime('%Y-%m-%d')
            
            # Parse string dates
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                try:
                    parsed_date = datetime.strptime(date_str, fmt)
                    return parsed_date.strftime('%Y-%m-%d')
                except ValueError:
                    continue
        except:
            pass
        
        return date_str  # Return as-is if parsing fails
    
    def _parse_datetime(self, datetime_str: str) -> Optional[datetime]:
        """Parse datetime string"""
        if not datetime_str:
            return None
        
        try:
            if isinstance(datetime_str, datetime):
                return datetime_str
            
            return datetime.fromisoformat(datetime_str.replace('Z', '+00:00'))
        except:
            try:
                return datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S')
            except:
                return None
    
    def get_export_summary(self) -> Dict:
        """
        Get summary of available export formats and recent exports
        
        Useful for providing users with information about export capabilities.
        """
        try:
            # Count files in exports directory
            export_files = list(self.config.EXPORTS_DIR.glob('*'))
            
            file_types = {}
            recent_files = []
            
            for file_path in export_files:
                if file_path.is_file():
                    ext = file_path.suffix.lower()
                    file_types[ext] = file_types.get(ext, 0) + 1
                    
                    # Get file info
                    stat = file_path.stat()
                    recent_files.append({
                        'name': file_path.name,
                        'size': stat.st_size,
                        'created': datetime.fromtimestamp(stat.st_ctime),
                        'type': ext
                    })
            
            # Sort by creation time, most recent first
            recent_files.sort(key=lambda x: x['created'], reverse=True)
            
            return {
                'supported_formats': ['Excel (.xlsx)', 'PDF Report (.pdf)', 'JSON (.json)', 'CSV (.csv)'],
                'export_directory': str(self.config.EXPORTS_DIR),
                'total_exports': len(export_files),
                'file_types': file_types,
                'recent_exports': recent_files[:10]  # Last 10 exports
            }
            
        except Exception as e:
            logger.error(f"Failed to get export summary: {str(e)}")
            return {}
    
    def cleanup_old_exports(self, days_old: int = 30) -> int:
        """
        Clean up export files older than specified days
        
        Helps manage disk space by removing old export files.
        """
        try:
            cutoff_date = datetime.now() - timedelta(days=days_old)
            removed_count = 0
            
            for file_path in self.config.EXPORTS_DIR.glob('*'):
                if file_path.is_file():
                    file_time = datetime.fromtimestamp(file_path.stat().st_ctime)
                    if file_time < cutoff_date:
                        file_path.unlink()
                        removed_count += 1
            
            logger.info(f"Cleaned up {removed_count} old export files")
            return removed_count
            
        except Exception as e:
            logger.error(f"Failed to cleanup exports: {str(e)}")
            return 0